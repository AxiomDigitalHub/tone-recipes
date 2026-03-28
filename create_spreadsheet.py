from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

wb = Workbook()

header_fill = PatternFill('solid', fgColor='2C3E50')
header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
data_font = Font(name='Arial', size=10)
alt_fill = PatternFill('solid', fgColor='F2F3F4')
white_fill = PatternFill('solid', fgColor='FFFFFF')
thin_border = Border(
    left=Side(style='thin', color='D5D8DC'),
    right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'),
    bottom=Side(style='thin', color='D5D8DC')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

topics = [
    [1, "Signal Chain Mastery", "The Complete Guide to Guitar Signal Chain Order \u2014 Why It Matters and How to Get It Right", "signal chain order guitar", "Informational", 9, 10, "Definitive guide", "FAQ + HowTo schema", "Top", "High", ""],
    [2, "Signal Chain Mastery", "Overdrive vs Distortion vs Fuzz: The Definitive Guide for Modeler Users", "overdrive vs distortion vs fuzz", "Informational", 9, 9, "Definitive guide", "FAQ schema", "Top", "High", ""],
    [3, "Signal Chain Mastery", "4-Cable Method Explained: How to Integrate Your Modeler with a Real Amp", "4 cable method guitar", "Informational", 8, 9, "How-to article", "HowTo schema", "Middle", "Medium", ""],
    [4, "Signal Chain Mastery", "Effects Loop Explained: What It Is and Why You Need One", "effects loop guitar explained", "Informational", 7, 9, "How-to article", "FAQ schema", "Top", "Medium", ""],
    [5, "Signal Chain Mastery", "Gain Staging for Guitar: How to Stack Drives Without Mud", "gain staging guitar", "Informational", 7, 8, "How-to article", "HowTo schema", "Middle", "Medium", ""],
    [6, "Signal Chain Mastery", "Wet/Dry Guitar Rigs: The Complete Routing Guide", "wet dry guitar rig", "Informational", 6, 7, "Definitive guide", "HowTo schema", "Middle", "High", ""],
    [7, "Signal Chain Mastery", "Why Fuzz Pedals Go Before the Wah (And Other Signal Chain Secrets)", "fuzz before wah why", "Informational", 6, 8, "How-to article", "FAQ schema", "Middle", "Low", ""],
    [8, "Signal Chain Mastery", "Buffer vs True Bypass: What Every Pedalboard Needs to Know", "buffer vs true bypass guitar", "Informational", 7, 8, "Comparison post", "FAQ schema", "Top", "Medium", ""],
    [9, "Signal Chain Mastery", "Pre-Amp vs Power Amp: What Each Contributes to Your Tone", "preamp vs power amp guitar", "Informational", 6, 9, "How-to article", "FAQ schema", "Top", "Medium", ""],
    [10, "Signal Chain Mastery", "How to Use Parallel Routing on Your Modeler for Better Tones", "parallel routing guitar modeler", "Informational", 5, 6, "How-to article", "HowTo schema", "Middle", "Medium", ""],
    [11, "Platform-Specific Guides", "Line 6 Helix: Complete Beginner's Guide to Building Your First Preset", "helix beginner guide", "Informational", 8, 7, "Definitive guide", "HowTo schema", "Top", "High", ""],
    [12, "Platform-Specific Guides", "Quad Cortex vs Helix vs Kemper: Which Modeler Is Right for You?", "quad cortex vs helix vs kemper", "Commercial", 9, 8, "Comparison post", "Article schema", "Middle", "High", ""],
    [13, "Platform-Specific Guides", "Boss Katana Hidden Amps: How to Unlock Sneaky Amps via Tone Studio", "boss katana hidden amps sneaky amps", "Informational", 8, 7, "How-to article", "HowTo schema", "Middle", "Medium", ""],
    [14, "Platform-Specific Guides", "Best Helix Amp Models Ranked: Our Top 15 for Every Genre", "best helix amp models", "Commercial", 9, 7, "Listicle", "Article schema", "Middle", "High", ""],
    [15, "Platform-Specific Guides", "Neural DSP Quad Cortex Tips: 10 Things You Didn't Know It Could Do", "quad cortex tips tricks", "Informational", 7, 6, "Listicle", "Article schema", "Middle", "Medium", ""],
    [16, "Platform-Specific Guides", "How to Use Helix Snapshots Like a Pro", "helix snapshots guide", "Informational", 7, 7, "How-to article", "HowTo schema", "Middle", "Medium", ""],
    [17, "Platform-Specific Guides", "Boss Katana vs Helix vs Quad Cortex: Best Modeler Under $1000?", "boss katana vs helix vs quad cortex", "Commercial", 8, 7, "Comparison post", "Article schema", "Middle", "Medium", ""],
    [18, "Platform-Specific Guides", "TONEX vs Helix vs Quad Cortex: AI Capture vs Traditional Modeling", "tonex vs helix vs quad cortex", "Commercial", 7, 8, "Comparison post", "FAQ schema", "Middle", "High", ""],
    [19, "Platform-Specific Guides", "How to Import IRs on Helix, Quad Cortex, and Kemper", "import IR helix quad cortex", "Informational", 6, 7, "How-to article", "HowTo schema", "Middle", "Medium", ""],
    [20, "Platform-Specific Guides", "Fractal FM3 vs Quad Cortex: Which Is Better for Live Use?", "fractal fm3 vs quad cortex live", "Commercial", 7, 6, "Comparison post", "Article schema", "Middle", "Medium", ""],
    [21, "Artist Tone Breakdowns", "How to Get Stevie Ray Vaughan's Tone on Helix", "srv tone helix", "Informational", 8, 6, "How-to article", "HowTo schema", "Middle", "Medium", ""],
    [22, "Artist Tone Breakdowns", "How to Get David Gilmour's Comfortably Numb Solo Tone", "david gilmour tone comfortably numb", "Informational", 8, 6, "How-to article", "HowTo schema", "Middle", "Medium", ""],
    [23, "Artist Tone Breakdowns", "How to Get The Edge's Delay Tone (Dotted Eighth Note Guide)", "the edge delay tone dotted eighth", "Informational", 7, 7, "How-to article", "HowTo schema", "Middle", "Medium", ""],
    [24, "Artist Tone Breakdowns", "How to Get John Mayer's Clean Blues Tone on Any Modeler", "john mayer tone modeler", "Informational", 7, 6, "How-to article", "HowTo schema", "Middle", "Medium", ""],
    [25, "Artist Tone Breakdowns", "How to Get Kurt Cobain's Grunge Tone (It's Simpler Than You Think)", "kurt cobain grunge tone", "Informational", 7, 7, "How-to article", "HowTo schema", "Middle", "Low", ""],
    [26, "Artist Tone Breakdowns", "How to Get Slash's Les Paul Through Marshall Tone", "slash tone les paul marshall", "Informational", 7, 6, "How-to article", "HowTo schema", "Middle", "Medium", ""],
    [27, "Artist Tone Breakdowns", "How to Get Eddie Van Halen's Brown Sound on a Modeler", "eddie van halen brown sound modeler", "Informational", 7, 6, "How-to article", "HowTo schema", "Middle", "Medium", ""],
    [28, "Artist Tone Breakdowns", "How to Get Tom Morello's Guitar Sounds (Kill Switch, Wah Tricks, and More)", "tom morello guitar sounds", "Informational", 6, 6, "How-to article", "HowTo schema", "Middle", "Medium", ""],
    [29, "Artist Tone Breakdowns", "5 Iconic Guitar Tones You Can Nail in Under 5 Minutes on Helix", "iconic guitar tones helix quick", "Informational", 7, 5, "Listicle", "Article schema", "Top", "Medium", ""],
    [30, "Artist Tone Breakdowns", "The 10 Most-Searched Guitar Tones (And How to Get Them)", "most searched guitar tones how to", "Informational", 8, 7, "Listicle", "Article schema", "Top", "High", ""],
    [31, "Amp & Cab Science", "Fender vs Marshall vs Vox: The 3 Amp Voices Every Guitarist Should Know", "fender vs marshall vs vox tone", "Informational", 8, 9, "Comparison post", "FAQ schema", "Top", "Medium", ""],
    [32, "Amp & Cab Science", "EL34 vs 6L6 Power Tubes: How They Shape Your Amp's Tone", "el34 vs 6l6 tone difference", "Informational", 6, 9, "Comparison post", "FAQ schema", "Middle", "Medium", ""],
    [33, "Amp & Cab Science", "What Is an Impulse Response (IR) and Why Does It Matter?", "what is impulse response guitar", "Informational", 7, 10, "How-to article", "FAQ schema", "Top", "Low", ""],
    [34, "Amp & Cab Science", "Open-Back vs Closed-Back Cabinets: How They Change Your Sound", "open back vs closed back cabinet guitar", "Informational", 6, 9, "Comparison post", "FAQ schema", "Top", "Low", ""],
    [35, "Amp & Cab Science", "Best Microphone Positions for Guitar Cab Recording (SM57 Placement Guide)", "sm57 placement guitar cab", "Informational", 6, 7, "How-to article", "HowTo schema", "Middle", "Medium", ""],
    [36, "Amp & Cab Science", "Speaker Shootout: Greenback vs Vintage 30 vs Jensen vs Celestion Blue", "greenback vs vintage 30 vs jensen", "Commercial", 7, 7, "Comparison post", "Article schema", "Middle", "Medium", ""],
    [37, "Amp & Cab Science", "Why Your Modeler Tone Sounds Fizzy (And How to Fix It)", "modeler tone fizzy how to fix", "Informational", 8, 8, "How-to article", "FAQ schema", "Middle", "Low", ""],
    [38, "Amp & Cab Science", "Class A vs Class AB Guitar Amps: What's the Actual Difference?", "class a vs class ab guitar amp", "Informational", 5, 9, "How-to article", "FAQ schema", "Top", "Medium", ""],
    [39, "Amp & Cab Science", "Master Volume vs Non-Master Volume Amps: The Tone Difference Explained", "master volume vs non master volume amp", "Informational", 5, 8, "How-to article", "FAQ schema", "Middle", "Medium", ""],
    [40, "Amp & Cab Science", "How IRs Work: The Science Behind Cabinet Simulation", "how impulse responses work guitar", "Informational", 5, 8, "Definitive guide", "Article schema", "Middle", "High", ""],
    [41, "Effects Deep Dives", "Chorus vs Flanger vs Phaser: The Modulation Effects Explained", "chorus vs flanger vs phaser guitar", "Informational", 8, 9, "Comparison post", "FAQ schema", "Top", "Medium", ""],
    [42, "Effects Deep Dives", "Delay Types Explained: Analog vs Digital vs Tape", "delay types analog digital tape guitar", "Informational", 7, 9, "How-to article", "FAQ schema", "Top", "Medium", ""],
    [43, "Effects Deep Dives", "How to Use Delay Like The Edge (Dotted Eighth Note Tutorial)", "dotted eighth note delay tutorial", "Informational", 7, 7, "How-to article", "HowTo schema", "Middle", "Medium", ""],
    [44, "Effects Deep Dives", "Reverb Types Explained: Spring vs Plate vs Hall vs Room", "reverb types spring plate hall room", "Informational", 7, 9, "How-to article", "FAQ schema", "Top", "Medium", ""],
    [45, "Effects Deep Dives", "Compressor Pedal Explained: What It Does and When to Use One", "compressor pedal explained guitar", "Informational", 7, 9, "How-to article", "FAQ schema", "Top", "Low", ""],
    [46, "Effects Deep Dives", "Tremolo vs Vibrato: The Difference Most Guitarists Get Wrong", "tremolo vs vibrato difference guitar", "Informational", 6, 9, "How-to article", "FAQ schema", "Top", "Low", ""],
    [47, "Effects Deep Dives", "How to Use a Noise Gate Without Killing Your Sustain", "noise gate guitar settings", "Informational", 6, 7, "How-to article", "HowTo schema", "Middle", "Low", ""],
    [48, "Effects Deep Dives", "Expression Pedal Assignments: 5 Creative Uses on Your Modeler", "expression pedal modeler uses", "Informational", 5, 5, "Listicle", "Article schema", "Middle", "Medium", ""],
    [49, "Effects Deep Dives", "EQ Pedal Placement: Before or After Distortion?", "eq pedal placement before after distortion", "Informational", 6, 8, "How-to article", "FAQ schema", "Middle", "Low", ""],
    [50, "Effects Deep Dives", "The Uni-Vibe Effect: What It Is and How Hendrix Used It", "uni-vibe effect hendrix", "Informational", 5, 7, "How-to article", "FAQ schema", "Middle", "Low", ""],
    [51, "Modeler Workflow & Live Use", "3 Beginner Signal Chains Every Guitarist Should Know", "beginner signal chain guitar", "Informational", 7, 7, "Listicle", "Article schema", "Top", "Low", ""],
    [52, "Modeler Workflow & Live Use", "How to Build a Worship Guitar Rig on a Helix", "worship guitar rig helix", "Informational", 7, 5, "How-to article", "HowTo schema", "Middle", "Medium", ""],
    [53, "Modeler Workflow & Live Use", "Going Direct to FOH: The Complete Guide for Modeler Users", "direct to foh guitar modeler", "Informational", 6, 7, "Definitive guide", "HowTo schema", "Middle", "High", ""],
    [54, "Modeler Workflow & Live Use", "In-Ear Monitors for Guitarists: Why You Sound Different (And How to Fix It)", "in ear monitors guitar tone", "Informational", 6, 6, "How-to article", "FAQ schema", "Middle", "Medium", ""],
    [55, "Modeler Workflow & Live Use", "How to A/B Your Modeler Tone Against the Original Recording", "a/b modeler tone original recording", "Informational", 5, 5, "How-to article", "HowTo schema", "Middle", "Low", ""],
    [56, "Modeler Workflow & Live Use", "Modeler Latency Explained: Does It Actually Affect Your Playing?", "modeler latency affect playing", "Informational", 6, 8, "How-to article", "FAQ schema", "Top", "Low", ""],
    [57, "Modeler Workflow & Live Use", "FRFR Speaker vs Guitar Cab for Modelers: Which Sounds Better?", "frfr vs guitar cab modeler", "Commercial", 7, 7, "Comparison post", "FAQ schema", "Middle", "Medium", ""],
    [58, "Modeler Workflow & Live Use", "How to Use Your Modeler as an Audio Interface for Recording", "modeler audio interface recording", "Informational", 6, 6, "How-to article", "HowTo schema", "Middle", "Low", ""],
    [59, "Modeler Workflow & Live Use", "Preset Organization: How to Structure Your Modeler for a Full Gig", "modeler preset organization gig", "Informational", 5, 5, "How-to article", "HowTo schema", "Middle", "Medium", ""],
    [60, "Modeler Workflow & Live Use", "Why Your Bedroom Tone Doesn't Work at Band Practice (Volume-Dependent EQ)", "bedroom tone band practice volume", "Informational", 7, 7, "How-to article", "FAQ schema", "Top", "Low", ""],
    [61, "Gear & Buying Guides", "Best Budget Modeler 2026: Under $500 Options Ranked", "best budget modeler 2026", "Commercial", 8, 6, "Listicle", "Article schema", "Bottom", "High", ""],
    [62, "Gear & Buying Guides", "What Is a Tone Recipe? How ToneRecipes Works", "what is tone recipe", "Informational", 4, 7, "How-to article", "FAQ schema", "Top", "Low", ""],
    [63, "Gear & Buying Guides", "Helix vs HX Stomp vs HX Effects: Which Line 6 Board Do You Need?", "helix vs hx stomp vs hx effects", "Commercial", 8, 7, "Comparison post", "FAQ schema", "Middle", "Medium", ""],
    [64, "Gear & Buying Guides", "Neural Amp Modeler (NAM) Explained: Free AI Amp Modeling", "neural amp modeler explained", "Informational", 7, 9, "How-to article", "FAQ schema", "Top", "Medium", ""],
    [65, "Gear & Buying Guides", "Kemper Profiler MK 2 vs Quad Cortex: The 2026 Flagship Showdown", "kemper mk2 vs quad cortex 2026", "Commercial", 8, 6, "Comparison post", "Article schema", "Middle", "High", ""],
    [66, "Gear & Buying Guides", "Do You Need a Power Amp with Your Modeler? (FRFR vs Power Amp + Cab)", "power amp modeler frfr", "Commercial", 6, 7, "How-to article", "FAQ schema", "Middle", "Medium", ""],
    [67, "Gear & Buying Guides", "Best IRs for Helix: Our Top Picks for Every Genre", "best IRs helix", "Commercial", 7, 5, "Listicle", "Article schema", "Middle", "Medium", ""],
    [68, "Gear & Buying Guides", "How Neural Capture Works on the Quad Cortex (And Why It Matters)", "neural capture quad cortex how it works", "Informational", 6, 8, "How-to article", "FAQ schema", "Top", "Medium", ""],
    [69, "Gear & Buying Guides", "Helix Stadium XL: Everything We Know About Line 6's New Flagship", "helix stadium xl review specs", "Commercial", 8, 6, "How-to article", "Article schema", "Middle", "Medium", ""],
    [70, "Gear & Buying Guides", "Best Modeler for Metal in 2026", "best modeler for metal 2026", "Commercial", 8, 6, "Listicle", "Article schema", "Bottom", "Medium", ""],
]

biz_relevance = {
    "Signal Chain Mastery": 7,
    "Platform-Specific Guides": 8,
    "Artist Tone Breakdowns": 7,
    "Amp & Cab Science": 7,
    "Effects Deep Dives": 7,
    "Modeler Workflow & Live Use": 7,
    "Gear & Buying Guides": 7,
}

def apply_header(ws, headers, col_widths):
    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'A2'

def style_data_cell(cell, row_idx, align=None):
    cell.font = data_font
    cell.fill = alt_fill if row_idx % 2 == 0 else white_fill
    cell.alignment = align or left_align
    cell.border = thin_border

def add_score_formatting(ws, col_letter, max_row):
    rng = f'{col_letter}2:{col_letter}{max_row}'
    ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThanOrEqual', formula=['8'], fill=PatternFill('solid', fgColor='C6EFCE'), font=Font(color='006100')))
    ws.conditional_formatting.add(rng, CellIsRule(operator='between', formula=['5', '7.9'], fill=PatternFill('solid', fgColor='FFEB9C'), font=Font(color='9C6500')))
    ws.conditional_formatting.add(rng, CellIsRule(operator='lessThan', formula=['5'], fill=PatternFill('solid', fgColor='FFC7CE'), font=Font(color='9C0006')))

# ===== SHEET 1: Topic List =====
ws1 = wb.active
ws1.title = "Topic List"

headers1 = ["Priority Rank", "Topic Cluster", "Topic Title", "Target Keyword", "Search Intent",
            "SEO Score", "AEO Score", "Priority Score", "Content Type", "AEO Format",
            "Funnel Stage", "Estimated Difficulty", "Notes"]
widths1 = [13, 24, 65, 38, 15, 11, 11, 13, 17, 22, 13, 18, 20]

for r, t in enumerate(topics, 2):
    rank, cluster, title, kw, intent, seo, aeo, ctype, aeo_fmt, funnel, diff, notes = t
    biz = 9 if rank == 62 else biz_relevance.get(cluster, 7)
    ps_formula = f'=F{r}*0.4+G{r}*0.3+{biz}*0.3'
    row_data = [rank, cluster, title, kw, intent, seo, aeo, ps_formula, ctype, aeo_fmt, funnel, diff, notes]
    for c, val in enumerate(row_data, 1):
        cell = ws1.cell(row=r, column=c, value=val)
        a = center_align if c not in (2, 3, 4, 13) else left_align
        style_data_cell(cell, r, a)
    ws1.cell(row=r, column=8).number_format = '0.0'

apply_header(ws1, headers1, widths1)
ws1.auto_filter.ref = f"A1:M{len(topics)+1}"

for col in ['F', 'G', 'H']:
    add_score_formatting(ws1, col, len(topics) + 1)

# ===== SHEET 2: Cluster Map =====
ws2 = wb.create_sheet("Cluster Map")

headers2 = ["Cluster Name", "Pillar Topic", "Topic Count", "Avg SEO Score", "Avg AEO Score",
            "Avg Priority Score", "Recommended Priority", "Notes"]
widths2 = [26, 65, 13, 15, 15, 17, 22, 50]

clusters_info = [
    ["Signal Chain Mastery", "The Complete Guide to Guitar Signal Chain Order \u2014 Why It Matters and How to Get It Right", 10, "1", "Highest combined SEO+AEO scores; foundational content"],
    ["Platform-Specific Guides", "Quad Cortex vs Helix vs Kemper: Which Modeler Is Right for You?", 10, "2", "High commercial intent; drives product-related traffic"],
    ["Artist Tone Breakdowns", "The 10 Most-Searched Guitar Tones (And How to Get Them)", 10, "3", "High search volume; strong engagement potential"],
    ["Amp & Cab Science", "Fender vs Marshall vs Vox: The 3 Amp Voices Every Guitarist Should Know", 10, "4", "Strong AEO scores; educational authority builder"],
    ["Effects Deep Dives", "Chorus vs Flanger vs Phaser: The Modulation Effects Explained", 10, "5", "Strong AEO; evergreen educational content"],
    ["Modeler Workflow & Live Use", "Going Direct to FOH: The Complete Guide for Modeler Users", 10, "6", "Practical utility; builds user loyalty"],
    ["Gear & Buying Guides", "Best Budget Modeler 2026: Under $500 Options Ranked", 10, "7", "Commercial intent but time-sensitive; refresh annually"],
]

cluster_ranges = [(2, 11), (12, 21), (22, 31), (32, 41), (42, 51), (52, 61), (62, 71)]

for r, (ci, cr) in enumerate(zip(clusters_info, cluster_ranges), 2):
    name, pillar, count, rec_pri, notes = ci
    start, end = cr
    avg_seo = f"=AVERAGE('Topic List'!F{start}:F{end})"
    avg_aeo = f"=AVERAGE('Topic List'!G{start}:G{end})"
    avg_pri = f"=AVERAGE('Topic List'!H{start}:H{end})"
    row_data = [name, pillar, count, avg_seo, avg_aeo, avg_pri, rec_pri, notes]
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        a = center_align if c in (3, 4, 5, 6, 7) else left_align
        style_data_cell(cell, r, a)
    ws2.cell(row=r, column=4).number_format = '0.0'
    ws2.cell(row=r, column=5).number_format = '0.0'
    ws2.cell(row=r, column=6).number_format = '0.0'

apply_header(ws2, headers2, widths2)
ws2.auto_filter.ref = f"A1:H{len(clusters_info)+1}"

for col in ['D', 'E', 'F']:
    add_score_formatting(ws2, col, len(clusters_info) + 1)

# ===== SHEET 3: AEO Quick Wins =====
ws3 = wb.create_sheet("AEO Quick Wins")

headers3 = ["Rank", "Topic Title", "Target Keyword", "AEO Score", "SEO Score",
            "AI Answer Format", "Schema Type", "Featured Snippet Opportunity",
            "AI Platform Opportunity", "Funnel Stage"]
widths3 = [8, 65, 38, 11, 11, 28, 22, 30, 30, 13]

ai_format_map = {
    "FAQ schema": "Direct Q&A paragraph",
    "FAQ + HowTo schema": "Direct Q&A + Step list",
    "HowTo schema": "Numbered step list",
    "Article schema": "Summary paragraph + list",
}
snippet_map = {
    "Comparison post": "Comparison table / list",
    "Definitive guide": "Definition box / paragraph",
    "How-to article": "Step-by-step list / paragraph",
    "Listicle": "Numbered list",
}
ai_platform_map = {
    "FAQ schema": "ChatGPT, Perplexity, Google SGE",
    "FAQ + HowTo schema": "ChatGPT, Perplexity, Google SGE",
    "HowTo schema": "Google SGE, Perplexity",
    "Article schema": "Perplexity, Google SGE",
}

sorted_topics = sorted(topics, key=lambda x: (-x[6], -x[5]))
top20 = sorted_topics[:20]

for r, t in enumerate(top20, 2):
    rank_num, cluster, title, kw, intent, seo, aeo, ctype, aeo_fmt, funnel, diff, notes = t
    ai_fmt = ai_format_map.get(aeo_fmt, "Summary paragraph")
    schema = aeo_fmt
    snippet = snippet_map.get(ctype, "Paragraph")
    ai_plat = ai_platform_map.get(aeo_fmt, "Perplexity, Google SGE")
    row_data = [r - 1, title, kw, aeo, seo, ai_fmt, schema, snippet, ai_plat, funnel]
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        a = center_align if c in (1, 4, 5, 10) else left_align
        style_data_cell(cell, r, a)

apply_header(ws3, headers3, widths3)
ws3.auto_filter.ref = f"A1:J{len(top20)+1}"

for col in ['D', 'E']:
    add_score_formatting(ws3, col, len(top20) + 1)

output_path = '/Users/daniellivengood/Documents/Claude/tone-recipes/content-topics.xlsx'
wb.save(output_path)
print(f"Saved to {output_path}")
